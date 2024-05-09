from tkinter import *
from tkinter import ttk, messagebox
from customtkinter import *
from PIL import Image
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
from main import run
from pathlib import Path


class UIHelpers:
    @staticmethod
    def create_frame(parent, row, column, padx=(0, 0), pady=(0, 0), rowspan=1, columnspan=1, sticky="nsew", fg_color="transparent"):
        frame = CTkFrame(parent, fg_color=fg_color)
        frame.grid(row=row, column=column, sticky=sticky, padx=padx, pady=pady, rowspan=rowspan, columnspan=columnspan)

        return frame

    @staticmethod
    def create_label(parent, text, font_size=48, bold=True, text_color="#C3C3C3", row=0, column=0, columnspan=1, anchor="w", padx=0, pady=0):
        font = ("Berlin Sans FB Demi", font_size, "bold" if bold else "normal")
        label = CTkLabel(parent, text=text, text_color=text_color, anchor=anchor, font=font)
        label.grid(row=row, column=column, columnspan=columnspan, sticky=anchor, padx=padx, pady=pady)
        return label

    @staticmethod
    def create_treeview(parent, columns, headings=None, height=8, pack=True):
        style = ttk.Style(parent)
        style.theme_use("clam")
        bg_color = "#2A2D2E"
        style.configure("Treeview",
                        background=bg_color,
                        fieldbackground=bg_color,
                        foreground="white",
                        rowheight=25,
                        bordercolor=bg_color,  # Border color
                        lightcolor=bg_color,  # Lighter lines in the tree
                        darkcolor=bg_color,  # Darker lines in the tree
                        )
        style.configure("Treeview.Heading",
                        background="#565B5E",
                        foreground="white",
                        font=("Helvetica", 10, "bold"),
                        bordercolor="#565B5E",
                        lightcolor="#565B5E",  # Lighter lines in the tree
                        darkcolor="#565B5E",
                        )
        style.configure(
            "Vertical.TScrollbar",
            background="#666666",  # Color of the scrollbar thumb
            troughcolor="#333333",  # Color of the scrollbar track
            relief="flat",  # Flat style
        )
        style.configure(
            "Horizontal.TScrollbar",
            background="#666666",
            troughcolor="#333333",
            relief="flat",
        )
        style.map(
            "Vertical.TScrollbar",
            background=[("active", "#666666"), ("disabled", "#666666")],  # Keep color consistent
            troughcolor=[("active", "#333333"), ("disabled", "#333333")],  # Same for trough
        )

        style.map(
            "Horizontal.TScrollbar",
            background=[("active", "#666666"), ("disabled", "#666666")],  # Consistent color
            troughcolor=[("active", "#333333"), ("disabled", "#333333")],  # Consistent trough
        )

        treeview = ttk.Treeview(parent, columns=columns, show="headings", height=height)
        if pack is True:
            treeview.pack(fill=BOTH, expand=True)

        if headings is None:
            headings = columns
        for idx, col in enumerate(columns):
            treeview.heading(col, text=headings[idx])
            treeview.column(col, width=100)
            if pack is False:
                treeview.column(col, width=100, stretch=False)

        return treeview


class Gui:
    def __init__(self, root, folder):
        self.root = root
        self.add_classes_func()
        self.add_new_status = folder
        self.settings_func()

        # none when want to create new
        # no none when want to open existing

        if folder is not None:
            self.folder = folder
            print(self.folder)
            self.add_settings(self.folder)

    def settings_func(self):
        if hasattr(self, "class_frame"):
            self.class_frame.pack_forget()

        self.settings_frame = CTkFrame(self.root, fg_color="transparent")
        self.settings_frame.pack(fill=BOTH, expand=1, padx=90, pady=35)
        self.settings_frame.columnconfigure((0, 1, 2), weight=1, uniform="a")
        self.settings_frame.rowconfigure((0, 1, 2), weight=1, uniform="a")
        self.settings_frame.rowconfigure(3, weight=3, uniform="a")

        lbl = CTkLabel(self.settings_frame, text="Settings", text_color="#C3C3C3", anchor=W, font=("Berlin Sans FB Demi", 48, "bold"))
        lbl.grid(row=0, column=0, sticky=W, padx=0, columnspan=2)

        self.Name = CTkEntry(self.settings_frame, placeholder_text="Project Name", width=520)
        self.Name.grid(row=1, column=0, sticky=NW, columnspan=2, ipadx=30, pady=25)

        self.days = CTkEntry(self.settings_frame, placeholder_text="Total days per week", width=420)
        self.days.grid(row=2, column=0, sticky=NW, columnspan=2, ipadx=30, pady=(0, 0))

        self.classes = CTkEntry(self.settings_frame, placeholder_text="Total classes per day", width=420)
        self.classes.grid(row=2, column=0, sticky=SW, columnspan=2, ipadx=30, pady=0)

        np_btn_frame = UIHelpers.create_frame(self.settings_frame,2, 0, sticky="w", columnspan=3, rowspan=2, pady=(0, 0), padx=(0,0))
        prev_btn = CTkButton(np_btn_frame, text="Prev", text_color="white", fg_color="transparent", border_width=2, hover_color="#2D2D2D", font=("Berlin Sans FB Demi", 20, "bold"), width=150, height=40, command=self.goto_dashboard)
        prev_btn.grid(row=0, column=0, sticky=W)
        next_btn = CTkButton(np_btn_frame, text="Next", text_color="white", fg_color="transparent", border_width=2, hover_color="#2D2D2D", font=("Berlin Sans FB Demi", 20, "bold"), width=150, height=40, command=self.settings_next)
        next_btn.grid(row=0, column=1, sticky=W, padx=(20, 0))

        if hasattr(self, "class_frame"):
            self.class_frame.pack_forget()
            try:
                self.add_settings(self.folder)
            except:
                pass

    def add_classes_func(self):
        if hasattr(self, "teachers_frame"):
            self.teachers_frame.pack_forget()
        if hasattr(self, "settings_frame"):
            self.settings_frame.pack_forget()

        print("hi")
        self.class_frame = CTkFrame(self.root, fg_color="transparent")
        self.class_frame.pack(fill=BOTH, expand=1, padx=90, pady=35)
        self.class_frame.columnconfigure((0, 1, 2, 3), weight=1, uniform="a")
        self.class_frame.rowconfigure((0, 1, 2, 4), weight=1, uniform="a")
        self.class_frame.rowconfigure(3, weight=3, uniform="a")

        lbl = CTkLabel(self.class_frame, text="ADD CLASSES", text_color="#C3C3C3", anchor=W, font=("Berlin Sans FB Demi", 48, "bold"))
        lbl.grid(row=0, column=0, sticky=W, padx=0, columnspan=2)

        upload_img = Image.open("Assets/upload.png")
        upload_img = CTkImage(upload_img)

        btn_frame = UIHelpers.create_frame(self.class_frame, 1, 0, columnspan=3, pady=10, sticky="w")
        add_btn = CTkButton(btn_frame, text="ADD", text_color="black", fg_color="#F8D15E", hover_color="#F8EB67", image=upload_img, font=("Berlin Sans FB Demi", 20, "bold"), width=370, height=40, command=self.add_btn_class)
        add_btn.grid(row=0, column=0, sticky=W, columnspan=1, ipadx=20)
        update_btn = CTkButton(btn_frame, text="Update", text_color="black", fg_color="#F8D15E", hover_color="#F8EB67", font=("Berlin Sans FB Demi", 20, "bold"), width=220, height=40, command=self.update_btn_class)
        update_btn.grid(row=0, column=1, sticky=W, columnspan=2, ipadx=20, padx=(20, 0))

        tree_frame = UIHelpers.create_frame(self.class_frame, 3, 0, columnspan=3, rowspan=2, pady=(0, 60), sticky="nsew")

        columns = ("class_name", "divisions")
        headings = ("Class Name", "Divisions")
        self.Class_Treeview = UIHelpers.create_treeview(tree_frame, columns, headings=headings, height=8)

        scroll_y = ttk.Scrollbar(self.class_frame, orient="vertical", style="Vertical.TScrollbar", command=self.Class_Treeview.yview)
        scroll_x = ttk.Scrollbar(tree_frame, orient=HORIZONTAL, style="Horizontal.TScrollbar", command=self.Class_Treeview.xview)
        self.Class_Treeview.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        scroll_x.pack(side=BOTTOM, fill=X, padx=(0, 15))
        scroll_y.grid(row=3, column=2, sticky="nse", columnspan=1, rowspan=2, pady=(28, 60), padx=(0, 2))

        np_btn_frame = UIHelpers.create_frame(self.class_frame,4, 2, sticky="w", columnspan=3, rowspan=2, pady=(100, 0), padx=(0,0))
        prev_btn = CTkButton(np_btn_frame, text="Prev", text_color="white", fg_color="transparent", border_width=2, hover_color="#2D2D2D", font=("Berlin Sans FB Demi", 20, "bold"), width=150, height=40, command=self.settings_func)
        prev_btn.grid(row=4, column=1, sticky=W)
        next_btn = CTkButton(np_btn_frame, text="Next", text_color="white", fg_color="transparent", border_width=2, hover_color="#2D2D2D", font=("Berlin Sans FB Demi", 20, "bold"), width=150, height=40, command=self.add_teachers_func)
        next_btn.grid(row=4, column=2, sticky=W, padx=(20, 0))

        if hasattr(self, "add_new_status"):
            if self.add_new_status is not None:
                self.update_btn_class()

    def add_teachers_func(self):
        if hasattr(self, "class_frame"):
            self.class_frame.pack_forget()
        if hasattr(self, "generate_frame"):
            self.generate_frame.pack_forget()
        if hasattr(self, "settings_frame"):
            self.settings_frame.pack_forget()

        self.teachers_frame = CTkFrame(self.root, fg_color="transparent")
        self.teachers_frame.pack(fill=BOTH, expand=1, padx=90, pady=35)
        self.teachers_frame.columnconfigure((0, 1, 2, 3), weight=1, uniform="a")
        self.teachers_frame.rowconfigure((0, 1, 2, 4), weight=1, uniform="a")
        self.teachers_frame.rowconfigure(3, weight=3, uniform="a")

        lbl = CTkLabel(self.teachers_frame, text="ADD TEACHERS", text_color="#C3C3C3", anchor=W, font=("Berlin Sans FB Demi", 48, "bold"))
        lbl.grid(row=0, column=0, sticky=W, padx=0, columnspan=2)

        upload_img = Image.open("Assets/upload.png")
        upload_img = CTkImage(upload_img)

        btn_frame = UIHelpers.create_frame(self.teachers_frame, 1, 0, columnspan=3, pady=10, sticky="w")
        add_btn = CTkButton(btn_frame, text="ADD", text_color="black", fg_color="#F8D15E", hover_color="#F8EB67", image=upload_img, font=("Berlin Sans FB Demi", 20, "bold"), width=370, height=40, command=self.add_btn_teachers)
        add_btn.grid(row=0, column=0, sticky=W, columnspan=1, ipadx=20)
        update_btn = CTkButton(btn_frame, text="Update", text_color="black", fg_color="#F8D15E", hover_color="#F8EB67", font=("Berlin Sans FB Demi", 20, "bold"), width=220, height=40, command=self.update_btn_teachers)
        update_btn.grid(row=0, column=1, sticky=W, columnspan=2, ipadx=20, padx=(20, 0))

        search_frame = CTkFrame(self.teachers_frame, fg_color="transparent")
        # search_frame.grid(row=2, column=0, sticky=W, columnspan=3, pady=10)
        searchbar = CTkEntry(search_frame, placeholder_text="Search for class...", width=520)
        searchbar.grid(row=0, column=0, sticky=NW, columnspan=2, ipadx=30, pady=0)
        search_btn = CTkButton(search_frame, text="Search", fg_color="transparent", hover_color="#2D2D2D", border_width=2, text_color=("gray10", "#DCE4EE"))
        search_btn.grid(row=0, column=3, sticky=NE, padx=(20, 0), pady=0, columnspan=3)

        tree_frame = UIHelpers.create_frame(self.teachers_frame, 3, 0, columnspan=3, rowspan=2, pady=(0, 60))

        columns = ("class_name", "divisions", "subject", "teacher", "frequency")
        headings = ("Class Name", "Divisions", "Subject", "Teacher", "Frequency")
        self.Teachers_Treeview = UIHelpers.create_treeview(tree_frame, columns, headings=headings, height=8)

        scroll_y = ttk.Scrollbar(self.teachers_frame, orient="vertical", style="Vertical.TScrollbar", command=self.Teachers_Treeview.yview)
        scroll_x = ttk.Scrollbar(tree_frame, orient=HORIZONTAL, style="Horizontal.TScrollbar", command=self.Teachers_Treeview.xview)
        self.Teachers_Treeview.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        scroll_x.pack(side=BOTTOM, fill=X, padx=(0, 15))
        scroll_y.grid(row=3, column=2, sticky="nse", columnspan=1, rowspan=2, pady=(28, 60), padx=(0, 2))

        np_btn_frame = UIHelpers.create_frame(self.teachers_frame,4, 2, sticky="w", columnspan=3, rowspan=2, pady=(100, 0), padx=(0,0))
        prev_btn = CTkButton(np_btn_frame, text="Prev", text_color="white", fg_color="transparent", border_width=2, hover_color="#2D2D2D", font=("Berlin Sans FB Demi", 20, "bold"), width=150, height=40, command=self.add_classes_func)
        prev_btn.grid(row=4, column=1, sticky=W)
        next_btn = CTkButton(np_btn_frame, text="Next", text_color="white", fg_color="transparent", border_width=2, hover_color="#2D2D2D", font=("Berlin Sans FB Demi", 20, "bold"), width=150, height=40, command=self.generate_func)
        next_btn.grid(row=4, column=2, sticky=W, padx=(20, 0))

        if hasattr(self, "add_new_status"):
            if self.add_new_status is not None:
                self.update_btn_teachers()

    def generate_func(self):
        if hasattr(self, "teachers_frame"):
            self.teachers_frame.pack_forget()

        self.generate_frame = CTkFrame(self.root, fg_color="transparent")
        self.generate_frame.pack(fill=BOTH, expand=1, padx=90, pady=35)
        self.generate_frame.columnconfigure((0, 1, 2, 3), weight=1, uniform="a")
        self.generate_frame.rowconfigure((0, 1, 2, 4), weight=1, uniform="a")
        self.generate_frame.rowconfigure(3, weight=3, uniform="a")

        lbl = CTkLabel(self.generate_frame, text="Generate", text_color="#C3C3C3", anchor=W, font=("Berlin Sans FB Demi", 48, "bold"))
        lbl.grid(row=0, column=0, sticky=W, padx=0, columnspan=2)

        upload_img = Image.open("Assets/download.png")
        upload_img = CTkImage(upload_img)

        btn_frame = UIHelpers.create_frame(self.generate_frame, 1, 0, columnspan=4, sticky="ew")
        gen_btn = CTkButton(btn_frame, text="Generate", text_color="black", fg_color="#F8D15E", hover_color="#F8EB67", font=("Berlin Sans FB Demi", 20, "bold"), width=200, height=40, command=self.generate_btn)
        gen_btn.grid(row=1, column=0, sticky=W, columnspan=2, ipadx=20, padx=(0, 250))
        download_btn = CTkButton(btn_frame, text="Download All", text_color="black", fg_color="#F8D15E", hover_color="#F8EB67", image=upload_img, font=("Berlin Sans FB Demi", 20, "bold"), width=300, height=40, command=self.download_btn)
        download_btn.grid(row=1, column=0, sticky=W, columnspan=2, ipadx=0, padx=(250, 0))
        finish_btn = CTkButton(btn_frame, text="Finish", text_color="black", fg_color="#F8D15E", hover_color="#F8EB67", font=("Berlin Sans FB Demi", 20, "bold"), width=200, height=40, command=self.goto_dashboard)
        finish_btn.grid(row=1, column=2, sticky=W, columnspan=3, ipadx=0, padx=(10, 0))
        prev_btn = CTkButton(btn_frame, text="Prev", text_color="white", fg_color="transparent", border_width=2, hover_color="#2D2D2D", font=("Berlin Sans FB Demi", 20, "bold"), width=150, height=40, command=self.add_teachers_func)
        prev_btn.grid(row=1, column=3, sticky=W, columnspan=2, padx=(220, 0))

        tab = CTkTabview(self.generate_frame, fg_color="#2A2D2E")
        tab.grid(row=2, column=0, sticky=NSEW, columnspan=3, rowspan=2, pady=(0, 0))

        tab_list = ["General", "Teachers", "Class Name"]
        tabs = {}
        for tab_name in tab_list:
            current_tab = tab.add(tab_name)
            tabs[tab_name] = current_tab
            current_tab.rowconfigure(0, weight=1)
            current_tab.columnconfigure(0, weight=1)

            label = CTkLabel(tabs[tab_name], text="Nothing To Download", text_color="#C3C3C3", anchor=W, font=("Berlin Sans FB Demi", 20))
            label.grid(row=0, column=0, sticky="nswe", padx=380)

        self.general_generate_tree = CTkScrollableFrame(tabs["General"], fg_color="transparent")
        self.general_generate_tree.grid(row=0, column=0, sticky="nswe", padx=10, pady=10)
        self.teachers_generate_tree = CTkScrollableFrame(tabs["Teachers"], fg_color="transparent")
        self.teachers_generate_tree.grid(row=0, column=0, sticky="nswe", padx=10, pady=10)
        self.class_generate_tree = CTkScrollableFrame(tabs["Class Name"], fg_color="transparent")
        self.class_generate_tree.grid(row=0, column=0, sticky="nswe", padx=10, pady=10)

        if hasattr(self, "add_new_status"):
            if self.add_new_status is not None:
                self.generate_btn()

    def goto_dashboard(self):
        for widget in self.root.winfo_children():  # Get all children in the root
            widget.destroy()
        Dashboard(self.root)

    def settings_next(self):
        def validate_input(inp):
            try:
                int(inp)
            except ValueError:
                return False
            return True

        self.folder = f"Timetables\\{self.Name.get()}"
        if self.Name.get() == "" or self.days.get() == "" or self.classes.get() == "":
            messagebox.showerror("Error", "Please enter all credentials")
        elif self.add_new_status is None and os.path.exists(self.folder):
            messagebox.showerror("Error", "ProjectName already exists")
        elif validate_input(self.days.get()) is False or validate_input(self.classes.get()) is False:
            messagebox.showerror("Error", "Please enter a valid integer")
        elif int(self.days.get()) > 7:
            messagebox.showerror("Error", "Days cannot be more than 7")
        else:
            if self.add_new_status is None:
                os.mkdir(self.folder)
                with open(f"Timetables\\{self.Name.get()}\\settings.txt", 'w') as file:
                    file.write(f"Days Per Week={self.days.get()}\n")  # Write 'x' value
                    file.write(f"Classes Per Day={self.classes.get()}\n")

            self.add_classes_func()

    def add_settings(self, name):
        path = Path(name)
        name = path.name
        self.Name.insert(0, name)
        self.Name.configure(state=DISABLED)

        with open(f"Timetables\\{name}\\settings.txt", 'r') as file:
            lines = file.readlines()

        for line in lines:
            key_value = line.strip().split('=')
            if len(key_value) == 2:  # Ensure there's a key and value
                key = key_value[0].strip()  # Get the variable name
                value = int(key_value[1].strip())  # Get the value and convert to integer

                if key == 'Days Per Week':
                    self.days.insert(0, value)
                elif key == 'Classes Per Day':
                    self.classes.insert(0, value)

    def add_btn_class(self):
        file_path = f"{self.folder}\\classes_added.xlsx"

        if not os.path.exists(file_path):
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet1"
            worksheet.cell(row=1, column=1, value="Classes")
            worksheet.cell(row=1, column=2, value="No of divisions")
            worksheet.cell(row=1, column=4, value="Note: save before you exit")
            workbook.save(file_path)
        os.startfile(file_path)

    def update_btn_class(self):
        file_path = f"{self.folder}\\classes_added.xlsx"
        df = pd.read_excel(file_path)
        first_two_columns = df.iloc[:, :2]
        self.Class_Treeview.delete(*self.Class_Treeview.get_children())
        for index, row in first_two_columns.iterrows():
            class_name = row["Classes"]
            divisions = row["No of divisions"]
            self.Class_Treeview.insert("", "end", values=(class_name, divisions))

    def add_btn_teachers(self):
        file_path = f"{self.folder}\\teachers_added.xlsx"
        df = pd.read_excel(f"{self.folder}\\classes_added.xlsx")
        class_names = df.iloc[:, :2]
        if not os.path.exists(file_path):
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Class Teachers"
            idx = 1
            for _, row in class_names.iterrows():
                class_name = row["Classes"]
                divisions = row["No of divisions"]
                worksheet.cell(row=1, column=idx, value="Class name")
                worksheet.cell(row=1, column=idx+1, value="Division")
                worksheet.cell(row=1, column=idx+2, value="Class Teacher name")

                for d in range(divisions):
                    worksheet.cell(row=d+2, column=idx, value=f"Class={class_name}")
                    worksheet.cell(row=d+2, column=idx+1, value=d+1)
                    # print(f"(row={d + 2}, column={idx}, value=class {class_name} division {d+1}")
                idx += 4
            for index, (_, row) in enumerate(class_names.iterrows()):
                class_name = row["Classes"]
                divisions = row["No of divisions"]
                worksheet = workbook.create_sheet(f"Class={class_name}")
                worksheet.cell(row=1, column=1, value="Subjects")
                worksheet.cell(row=1, column=2, value="Frequency per week")
                for div in range(int(divisions)):
                    worksheet.cell(row=1, column=3 + div, value=f"Teacher for division {div+1}")
            workbook.save(file_path)
        os.startfile(file_path)

    def update_btn_teachers(self):
        file_path = f"{self.folder}\\teachers_added.xlsx"
        df = pd.read_excel(file_path, sheet_name=None)
        self.data = {}
        self.Teachers_Treeview.delete(*self.Teachers_Treeview.get_children())

        for idx, (sheet_name, data_frame) in enumerate(df.items()):
            if idx > 0:
                teacher_start_index = 2
                num_teacher_columns = len(data_frame.columns) - teacher_start_index

                for _, row in data_frame.iterrows():
                    subject = row.iloc[0]  # First column
                    frequency = row.iloc[1]  # Second column

                    # Iterate over teacher columns by index
                    for teacher_idx in range(num_teacher_columns):
                        teacher_name = row.iloc[teacher_start_index + teacher_idx]  # Access by index
                        # print(f"{subject}\t{frequency}\t{teacher_name}\t{teacher_idx+1}\t{sheet_name}")

                        # convert sheet_name to lower case
                        key = f"{sheet_name.split('=')[1]}{chr(teacher_idx+1+64)}"
                        if key not in self.data:
                            self.data[key] = [[subject], [teacher_name], [frequency]]
                        else:
                            self.data[key][0].append(subject)
                            self.data[key][1].append(teacher_name)
                            self.data[key][2].append(frequency)
                        self.Teachers_Treeview.insert("", "end", values=(sheet_name.split('=')[1], chr(teacher_idx+1+64), subject, teacher_name, frequency))
            else:
                df = data_frame
                num_columns_per_group = 3
                for idx, row in df.iterrows():
                    for group_start in range(0, len(row), num_columns_per_group+1):
                        class_name = row.iloc[group_start]  # Class name
                        division = row.iloc[group_start + 1]  # Division
                        teacher_name = row.iloc[group_start + 2]  # Class Teacher name
                        try:
                            class_name = class_name.split('=')[1]
                        except:
                            pass
                        key = f"{class_name}{chr(division+64)}"
                        self.data[key] = [[], [], [], teacher_name]

    def generate_btn(self):
        global column_initials
        try:
            matrix, teacher_timetables, class_timetables = run(self.data, int(self.classes.get()), int(self.days.get()))
            self.general_tree_list = []
            self.teachers_tree_list = []
            self.classes_tree_list = []
            self.general_lbl_list = []
            self.teachers_lbl_list = []
            self.classes_lbl_list = []
            for day_idx, day in enumerate(matrix):
                day_name = f"Day {day_idx + 1}"

                # Create a Frame for each day's schedule
                day_frame = UIHelpers.create_frame(self.general_generate_tree, row=day_idx, column=0, pady=(10, 20))

                day_label = CTkLabel(day_frame, text=day_name, font=("Berlin Sans FB Demi", 16, "bold"))
                day_label.grid(row=0, column=0, columnspan=8, sticky="w", pady=(0, 10))

                # Determine the number of periods for column names
                num_periods = len(list(day.values())[0])  # Number of periods in the first class

                # Create columns for the Treeview
                column_initials = [f"Period {i + 1}" for i in range(num_periods)]
                columns = ["Class Name"] + column_initials

                # Create a Treeview for the day's schedule
                treeview = UIHelpers.create_treeview(day_frame, columns, headings=columns, pack=False)
                treeview.grid(row=1, column=0, sticky="nswe")
                self.general_tree_list.append(treeview)
                self.general_lbl_list.append(day_name)

                vsb = ttk.Scrollbar(day_frame, orient="vertical", command=treeview.yview)
                vsb.grid(row=1, column=1, sticky="ns")  # Place vertical scrollbar
                hsb = ttk.Scrollbar(day_frame, orient="horizontal", command=treeview.xview)
                hsb.grid(row=2, column=0, sticky="ew")  # Place horizontal scrollbar
                treeview.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

                # Populate the Treeview with the class schedules for the day
                for class_name, schedules in day.items():
                    row_data = [class_name]  # Start with the class name
                    row_data.extend([f"{sched[1]} - {sched[0]}" for sched in schedules])  # Add subjects and teachers
                    treeview.insert("", "end", values=row_data)  # Insert into the Treeview

                num_rows = len(treeview.get_children())
                min_height = 2
                treeview["height"] = max(num_rows, min_height)

            # --------------------Teachers time table------------------

            for teachers_idx, (teacher, timetable) in enumerate(teacher_timetables.items()):
                teacher_frame = UIHelpers.create_frame(self.teachers_generate_tree, row=teachers_idx, column=0, pady=(10, 10))

                # Add a label with the teacher's name
                teacher_label = CTkLabel(teacher_frame, text=f"Timetable for {teacher}", font=("Berlin Sans FB Demi", 16, "bold"))
                teacher_label.grid(row=0, column=0, columnspan=2, sticky="w")

                # Create the TreeView for the teacher's timetable
                columns2 = ["Day"] + column_initials

                treeview2 = UIHelpers.create_treeview(teacher_frame, columns2, headings=columns2, pack=False)
                treeview2.grid(row=1, column=0, sticky="nswe")
                self.teachers_tree_list.append(treeview2)
                self.teachers_lbl_list.append(teacher)

                # Add scrollbars to the TreeView
                vsb2 = ttk.Scrollbar(teacher_frame, orient="vertical", command=treeview2.yview)
                hsb2 = ttk.Scrollbar(teacher_frame, orient="horizontal", command=treeview2.xview)
                vsb2.grid(row=1, column=1, sticky="ns")
                hsb2.grid(row=2, column=0, sticky="ew")
                treeview2.configure(yscrollcommand=vsb2.set, xscrollcommand=hsb2.set)

                # Populate the TreeView with the teacher's timetable
                for day_idx, (day, schedule) in enumerate(timetable.items()):
                    row_data2 = [day]
                    for sched in schedule:
                        if sched != 0:
                            row_data2.append(f"{sched[1]} - {sched[0]}")
                        else:
                            row_data2.append("Free")
                    treeview2.insert("", "end", values=row_data2)

                num_rows = len(treeview2.get_children())
                min_height = 2
                treeview2["height"] = max(num_rows, min_height)

            # --------------------Class time table------------------
            for class_idx, (class_name, timetable) in enumerate(class_timetables.items()):
                class_frame = UIHelpers.create_frame(self.class_generate_tree, row=class_idx, column=0,pady=(10, 10))

                # Add a label with the teacher's name
                class_label = CTkLabel(class_frame, text=f"Timetable for {class_name}",
                                             font=("Berlin Sans FB Demi", 16, "bold"))
                class_label.grid(row=0, column=0, columnspan=2, sticky="w")

                # Create the TreeView for the teacher's timetable
                columns3 = ["Day"] + column_initials

                treeview3 = UIHelpers.create_treeview(class_frame, columns3, headings=columns3, pack=False)
                treeview3.grid(row=1, column=0, sticky="nswe")
                self.classes_tree_list.append(treeview3)
                self.classes_lbl_list.append(class_name)

                # Add scrollbars to the TreeView
                vsb3 = ttk.Scrollbar(class_frame, orient="vertical", command=treeview3.yview)
                hsb3 = ttk.Scrollbar(class_frame, orient="horizontal", command=treeview3.xview)
                vsb3.grid(row=1, column=1, sticky="ns")
                hsb3.grid(row=2, column=0, sticky="ew")
                treeview3.configure(yscrollcommand=vsb3.set, xscrollcommand=hsb3.set)

                # Populate the TreeView with the teacher's timetable
                for day_idx, (day, schedule) in enumerate(timetable.items()):
                    row_data3 = [day]
                    for sched in schedule:
                        if sched != 0:
                            row_data3.append(f"{sched[1]} - {sched[0]}")
                        else:
                            row_data3.append("Free")
                    treeview3.insert("", "end", values=row_data3)

                num_rows = len(treeview3.get_children())
                min_height = 2
                treeview3["height"] = max(num_rows, min_height)

            # print(matrix)
            # for mat in matrix:
            #     for key, dt in mat.items():
            #         print(key, dt)
            #     print()
            # print()
            # print()

            # for teacher_name, timetable in teacher_timetables.items():
            #     print(f"Timetable for {teacher_name}:")
            #     for day, periods in timetable.items():
            #         print(f"{day}: {periods}")
            #     print()  # New line for separation

            # for class_name, schedule in class_timetables.items():
            #     print(f"Timetable for {class_name}:")
            #     for day, lessons in schedule.items():
            #         print(f"  {day}: {lessons}")
        except Exception as e:
            print("An error occurred:", str(e))

    def export_treeview_to_excel(self, treeview_list, label_list, file_path, sheet_name):
        # Check if the workbook exists, and load it; otherwise, create a new one
        new_workbook = False
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)  # Load existing workbook
        else:
            workbook = Workbook()  # Create a new workbook
            new_workbook = True  # Indicates a new workbook was created

        # If it's the first sheet and there's an existing default 'Sheet', remove it
        if new_workbook and workbook.sheetnames[0] == "Sheet":
            del workbook["Sheet"]  # Remove default sheet if it exists

        # If the sheet already exists, use it; otherwise, create a new sheet
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]  # Use existing sheet
        else:
            worksheet = workbook.create_sheet(sheet_name)  # Create new sheet

        row_offset = worksheet.max_row + 1  # Start from the last used row

        for idx, (treeview, name) in enumerate(zip(treeview_list, label_list)):
            worksheet.cell(row=row_offset, column=1, value=name)  # Add name in first column

            # Get column headings from the Treeview
            column_headings = [treeview.heading(col)["text"] for col in treeview["columns"]]

            # Write the column headings to the worksheet
            for col_idx, heading in enumerate(column_headings, 2):
                worksheet.cell(row=row_offset, column=col_idx, value=heading)

            # Write the data from the Treeview
            for item in treeview.get_children():
                row_offset += 1
                values = treeview.item(item, "values")
                for col_idx, value in enumerate(values, 2):
                    worksheet.cell(row=row_offset, column=col_idx, value=value)

            # Add a blank row after each section for better separation
            row_offset += 1  # Leave one row blank at the end of each section

        # Save the workbook
        workbook.save(file_path)

    def download_btn(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

        if path is not None:
            self.export_treeview_to_excel(self.general_tree_list, self.general_lbl_list, path, "General")
            self.export_treeview_to_excel(self.teachers_tree_list, self.teachers_lbl_list, path, "Teachers")
            self.export_treeview_to_excel(self.classes_tree_list, self.classes_lbl_list, path, "Classes")


class Dashboard:
    def __init__(self, root):
        self.root = root
        self.root.title("Generate Time-Table")
        self.root.geometry("1200x600")
        set_appearance_mode("dark")
        set_default_color_theme("dark-blue")
        self.gui()

    def gui(self):
        self.root.columnconfigure((0, 1, 2), weight=1, uniform="a")  # Uniform and equal expansion for columns
        self.root.rowconfigure(2, weight=1)  # Row 2 expands
        self.root.rowconfigure(3, weight=1)  # Row 3 expands as well

        # Create the heading label
        heading_label = CTkLabel(self.root, text="Time Table", font=("Berlin Sans FB Demi", 50, "bold"))
        heading_label.grid(row=0, column=0, columnspan=3, pady=(20, 10), sticky='n')  # Centered at the top

        # Search frame and its components
        search_frame = CTkFrame(self.root, fg_color="transparent")
        search_frame.grid(row=1, column=1, columnspan=4, sticky=W, pady=10, padx=(70, 0))

        self.searchbar = CTkEntry(search_frame, placeholder_text="Search for time tables...", width=520)
        self.searchbar.grid(row=0, column=0, sticky=NW, columnspan=2, ipadx=30, pady=0)

        search_btn = CTkButton(search_frame, text="Search", fg_color="transparent", hover_color="#2D2D2D", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.search)
        search_btn.grid(row=0, column=3, sticky=NE, padx=(20, 0), pady=0, columnspan=3)

        show_all_btn = CTkButton(search_frame, text="Show All", fg_color="transparent", hover_color="#2D2D2D", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.show_all)
        show_all_btn.grid(row=0, column=5, sticky=NE, padx=(190, 0), pady=0, columnspan=3)

        # The view frame
        self.view_frame = CTkScrollableFrame(self.root)
        self.view_frame.grid(row=2, column=0, columnspan=3, rowspan=2, sticky='nsew', padx=100, pady=10)  # Spanning rows 2-3, expanding
        self.view_frame.columnconfigure((0, 1, 2, 3), weight=1, uniform="a")
        self.view_frame.rowconfigure((0, 1, 2), weight=1, uniform="a")

        self.dir = Path.cwd()
        self.dir = self.dir / "Timetables"
        self.folder_cache = [item.name for item in self.dir.iterdir() if item.is_dir()]
        self.show_all()

    def show_all(self):
        for widget in self.view_frame.winfo_children():  # Get all children in the root
            widget.destroy()
        add_img = Image.open("Assets/add_new.png")

        add_img = CTkImage(add_img)
        add_new_btn = CTkButton(self.view_frame, image=add_img, compound="bottom", text="Add New", font=("Berlin Sans FB Demi", 25, "bold"), corner_radius=10, fg_color="#454545", text_color="#C3C3C3", hover_color="#3F3F3F", width=210, height=200, command=self.generate_timetable)
        add_new_btn.grid(row=0, column=0, padx=120, pady=20, columnspan=2, sticky='w')

        for idx, folder in enumerate(self.folder_cache):
            btn = CTkButton(self.view_frame, text=folder, font=("Berlin Sans FB Demi", 25, "bold"), corner_radius=10, fg_color="#454545",
                                    text_color="#C3C3C3", hover_color="#3F3F3F", width=210, height=200,
                                    command=lambda f=folder: self.view_timetable(f))
            row = (idx+1) // 3
            col = (idx+1) % 3
            if len(self.folder_cache) > 1:
                btn.grid(row=row, column=col, padx=120, pady=20, columnspan=2, sticky='w')
            else:
                btn.grid(row=0, column=1, padx=350, pady=20, columnspan=2, sticky='w')

    def search(self):
        search = self.searchbar.get()
        if search != "":
            for widget in self.view_frame.winfo_children():  # Get all children in the root
                widget.destroy()
            self.dir = Path.cwd()
            self.dir = self.dir / "Timetables"
            folders = []
            for item in self.dir.iterdir():
                if item.is_dir() and search.lower() in item.name.lower():
                    folders.append(item)

            add_img = Image.open("Assets/add_new.png")

            add_img = CTkImage(add_img)
            add_new_btn = CTkButton(self.view_frame, image=add_img, compound="bottom", text="Add New",
                                    font=("Berlin Sans FB Demi", 25, "bold"), corner_radius=10, fg_color="#454545",
                                    text_color="#C3C3C3", hover_color="#3F3F3F", width=210, height=200,
                                    command=self.generate_timetable)
            add_new_btn.grid(row=0, column=0, padx=120, pady=20, columnspan=2, sticky='w')

            for idx, folder in enumerate(folders):
                folder = folder.name
                btn = CTkButton(self.view_frame, text=folder, font=("Berlin Sans FB Demi", 25, "bold"), corner_radius=10, fg_color="#454545",
                                        text_color="#C3C3C3", hover_color="#3F3F3F", width=210, height=200,
                                        command=lambda: self.view_timetable(folder))
                row = (idx+1)//3
                col = (idx+1)%3
                if len(folders) > 1:
                    btn.grid(row=row, column=col, padx=120, pady=20, columnspan=2, sticky='w')
                else:
                    btn.grid(row=0, column=1, padx=350, pady=20, columnspan=2, sticky='w')

    def generate_timetable(self):
        for widget in self.root.winfo_children():  # Get all children in the root
            widget.destroy()
        Gui(self.root, None)

    def view_timetable(self, folder):
        for widget in self.root.winfo_children():  # Get all children in the root
            widget.destroy()
        Gui(self.root, f"{self.dir}\\{folder}")


root = CTk()
Dashboard(root)
root.mainloop()
